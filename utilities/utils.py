import csv

import softest
from openpyxl import Workbook, load_workbook


class Utils(softest.TestCase):
    def assertListItemText(self, list1, value):
        for item in list1:
            print("The text is: " + item.text)
            self.soft_assert(self.assertEqual, item.text, value)
            if item.text == value:
                print("Test Passed")
            else:
                print("Test Failed")

            self.assert_all()

    def read_data_from_excel(fileName, sheet):
        dataList = []
        wb = load_workbook(filename=fileName)
        sh = wb[sheet]
        row_ct = sh.max_row
        column_ct = sh.max_column
        for i in range(2, row_ct + 1):
            row = []
            for j in range(1, column_ct + 1):
                row.append(sh.cell(row=i, column=j).value)
            dataList.append(row)
        return dataList

    def read_data_from_csv(fileName):
        dataList = []
        csvFile = open(fileName, "r")
        csvReader = csv.reader(csvFile)
        next(csvReader)
        for row in csvReader:
            dataList.append(row)

        return dataList
